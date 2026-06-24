import sys
import re
import openpyxl
from bs4 import BeautifulSoup, NavigableString, Tag

SKIP_TAGS = {'script', 'style'}
PURE_NUMBER_RE = re.compile(r'^\s*-?\d+(\.\d+)?\s*$')

EMOJI_RE = re.compile(
    r'[\U0001F000-\U0001FFFF'
    r'\U00002600-\U000027FF'
    r'←-⇿'
    r'⌀-⏿'
    r'■-◿'
    r'✀-➿'
    r']+'
)

_LEAD_PAT = re.compile(
    r'^([\s\U0001F000-\U0001FFFF\U00002600-\U000027FF←-⇿⌀-⏿■-◿✀-➿]*)'
)
_TRAIL_PAT = re.compile(
    r'([\s\U0001F000-\U0001FFFF\U00002600-\U000027FF←-⇿⌀-⏿■-◿✀-➿]*)$'
)


def merge_emoji(original, translated):
    lead_m = _LEAD_PAT.match(original)
    trail_m = _TRAIL_PAT.search(original)
    leading = lead_m.group().strip() if lead_m and EMOJI_RE.search(lead_m.group()) else ''
    trailing = trail_m.group().strip() if trail_m and EMOJI_RE.search(trail_m.group()) else ''
    result = translated.strip()
    if leading:
        result = leading + ' ' + result
    if trailing and trailing != leading:
        result = result + ' ' + trailing
    return result


def should_extract(text):
    text = text.strip()
    if not text:
        return False
    if PURE_NUMBER_RE.match(text):
        return False
    latin_letters = re.findall(r'[a-zA-Z]', text)
    if len(latin_letters) < 3:
        return False
    return True


# ─── NARR / VO helpers ───────────────────────────────────────────────────────

def _find_narr_script_offset(html_content: str) -> tuple | None:
    for m in re.finditer(r'<script[^>]*>(.*?)</script>', html_content,
                         re.DOTALL | re.IGNORECASE):
        if re.search(r'\bNARR\b', m.group(1)):
            return m.group(1), m.start(1)
    return None


def _extract_narr_literal_and_offset(script_text: str) -> tuple | None:
    m = re.search(r'\bNARR\s*=\s*\{', script_text)
    if not m:
        return None

    start = m.end() - 1
    depth = 0
    in_string = False
    string_char = None
    i = start

    while i < len(script_text):
        ch = script_text[i]
        if in_string:
            if ch == '\\' and i + 1 < len(script_text):
                i += 2
                continue
            if ch == string_char:
                in_string = False
        else:
            if ch in ('"', "'", '`'):
                in_string = True
                string_char = ch
            elif ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    return script_text[start:i + 1], start
        i += 1

    return None


def _js_escape(text: str, quote_char: str) -> str:
    text = text.replace('\\', '\\\\')
    text = text.replace(quote_char, '\\' + quote_char)
    return text


def _rebuild_narr_literal(narr_literal, translations, english_check):
    output = []
    value_idx = 0
    applied = 0
    i = 0
    n = len(narr_literal)

    while i < n:
        ch = narr_literal[i]

        if ch in ('"', "'", '`'):
            quote_char = ch
            end = i + 1
            while end < n:
                c = narr_literal[end]
                if c == '\\' and end + 1 < n:
                    end += 2
                    continue
                if c == quote_char:
                    break
                end += 1

            j = end + 1
            while j < n and narr_literal[j] in ' \t\n\r':
                j += 1
            is_key = j < n and narr_literal[j] == ':'

            if is_key:
                output.append(narr_literal[i:end + 1])
            else:
                if value_idx in translations:
                    telugu = translations[value_idx]
                    escaped = _js_escape(telugu, quote_char)
                    output.append(f'{quote_char}{escaped}{quote_char}')
                    applied += 1
                else:
                    output.append(narr_literal[i:end + 1])
                value_idx += 1

            i = end + 1
            continue

        output.append(ch)
        i += 1

    return ''.join(output), applied


def apply_narr_translations(html_content, vo_translations, vo_english_check):
    result = _find_narr_script_offset(html_content)
    if not result:
        return html_content, 0, 0
    script_text, script_offset = result

    narr_result = _extract_narr_literal_and_offset(script_text)
    if not narr_result:
        return html_content, 0, 0
    narr_literal, narr_start_in_script = narr_result

    new_narr, applied = _rebuild_narr_literal(narr_literal, vo_translations, vo_english_check)

    abs_start = script_offset + narr_start_in_script
    abs_end   = abs_start + len(narr_literal)
    new_html  = html_content[:abs_start] + new_narr + html_content[abs_end:]
    return new_html, applied, len(vo_english_check)


_VOICE_LANG_PATCHES = [
    (
        '/* lock to Google Hindi (hi-IN) */',
        '/* lock to Google Telugu (te-IN) */',
    ),
    (
        'if (lang.startsWith("hi-in")) s += 100;',
        'if (lang.startsWith("te-in")) s += 100;',
    ),
    (
        'else if (lang.startsWith("hi")) s += 90;',
        'else if (lang.startsWith("te")) s += 90;',
    ),
]

_LANG_FALLBACK_OLD = '} /* else: let the browser use its default LOCAL voice (reliable) */'
_LANG_FALLBACK_NEW = ("} else {\n"
                      "              u.lang = 'te-IN';\n"
                      "            }")


def update_voice_language(html_content):
    for old, new in _VOICE_LANG_PATCHES:
        html_content = html_content.replace(old, new, 1)
    if "u.lang = 'te-IN'" not in html_content:
        html_content = html_content.replace(_LANG_FALLBACK_OLD, _LANG_FALLBACK_NEW, 1)
    return html_content


# ─── Excel loader ────────────────────────────────────────────────────────────

def load_translations(xlsx_path):
    """
    Returns (ost_trans, ost_check, vo_trans, vo_check).
    Backward compatible: old single-sheet Excel returns empty VO dicts.
    """
    wb = openpyxl.load_workbook(xlsx_path)

    if 'On-Screen Text' in wb.sheetnames:
        ost_ws = wb['On-Screen Text']
    elif 'Translate Here' in wb.sheetnames:
        ost_ws = wb['Translate Here']
    else:
        ost_ws = wb.active

    ost_trans, ost_check = {}, {}
    for row in ost_ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        idx = int(row[0])
        ost_check[idx] = str(row[3] or '').strip()
        telugu = row[5]
        if telugu and str(telugu).strip():
            ost_trans[idx] = str(telugu).strip()

    vo_trans, vo_check = {}, {}
    if 'Narration VO' in wb.sheetnames:
        vo_ws = wb['Narration VO']
        for row in vo_ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            idx      = int(row[0])
            original = str(row[1] or '').strip()
            telugu   = row[3]
            vo_check[idx] = original
            if telugu and str(telugu).strip():
                vo_trans[idx] = str(telugu).strip()

    return ost_trans, ost_check, vo_trans, vo_check


# ─── OST DOM replacement (unchanged logic) ───────────────────────────────────

def collect_replacements(soup, translations, english_check):
    replacements = []
    index = 0
    seen_attr_ids = set()
    warnings = 0

    for node in soup.descendants:
        if isinstance(node, Tag):
            if node.has_attr('data-script'):
                node_id = node.get('id', '')
                if node_id not in seen_attr_ids:
                    text = node['data-script'].strip()
                    if should_extract(text):
                        seen_attr_ids.add(node_id)
                        if index in translations:
                            expected = english_check.get(index, '')
                            if expected and expected != text:
                                print(f'WARNING row {index + 3}: expected "{expected[:60]}" but found "{text[:60]}" — skipping')
                                warnings += 1
                            else:
                                final = merge_emoji(expected, translations[index])
                                replacements.append((node, final, 'ATTR'))
                        index += 1

        elif isinstance(node, NavigableString) and type(node) is NavigableString:
            if any(p.name in SKIP_TAGS for p in node.parents if isinstance(p, Tag)):
                continue
            text = str(node).strip()
            if should_extract(text):
                if index in translations:
                    expected = english_check.get(index, '')
                    if expected and expected != text:
                        print(f'WARNING row {index + 3}: expected "{expected[:60]}" but found "{text[:60]}" — skipping')
                        warnings += 1
                    else:
                        final = merge_emoji(expected, translations[index])
                        replacements.append((node, final, 'TEXT'))
                index += 1

    return replacements, warnings


def apply_replacements(replacements):
    for node, new_text, node_type in replacements:
        if node_type == 'ATTR':
            node['data-script'] = new_text
        else:
            node.replace_with(NavigableString(new_text))


def main():
    if len(sys.argv) != 4:
        print('Usage: python apply_translation.py "<input.html>" <translations.xlsx> "<output.html>"')
        sys.exit(1)

    html_path = sys.argv[1]
    xlsx_path = sys.argv[2]
    output_path = sys.argv[3]

    ost_trans, ost_check, vo_trans, vo_check = load_translations(xlsx_path)
    if not ost_trans and not vo_trans:
        print('No translations found in Excel file. Nothing to apply.')
        sys.exit(0)

    with open(html_path, encoding='utf-8') as f:
        html = f.read()

    # Apply OST via BeautifulSoup DOM manipulation
    soup = BeautifulSoup(html, 'html.parser')
    replacements, warnings = collect_replacements(soup, ost_trans, ost_check)
    apply_replacements(replacements)
    new_html = str(soup)

    # Apply VO narration on the raw string (avoids BS4 script-content encoding issues)
    if vo_trans:
        new_html, vo_applied, vo_total = apply_narr_translations(new_html, vo_trans, vo_check)
        blank_vo = vo_total - vo_applied
        print(f'VO: Applied {vo_applied}/{vo_total} narration translations ({blank_vo} left blank — originals kept)')

    # Patch voice language to te-IN
    new_html = update_voice_language(new_html)

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(new_html)

    ost_blank = len(ost_check) - len(replacements)
    print(f'OST: Applied {len(replacements)}/{len(ost_check)} translations ({ost_blank} rows left blank — originals kept)')
    if warnings:
        print(f'{warnings} drift warnings — those rows were skipped')
    print(f'Output written to {output_path}')


if __name__ == '__main__':
    main()

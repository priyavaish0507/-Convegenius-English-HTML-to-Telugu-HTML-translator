import sys
import time
import requests
import openpyxl

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

API_URL = 'https://api.sarvam.ai/translate'
MAX_CHARS = 1000   # mayura:v1 limit


def translate(text, api_key):
    """Translate a single string from English to Telugu via Sarvam AI."""
    payload = {
        'input': text,
        'source_language_code': 'en-IN',
        'target_language_code': 'te-IN',
        'model': 'mayura:v1',
        'mode': 'formal',
    }
    headers = {
        'api-subscription-key': api_key,
        'Content-Type': 'application/json',
    }
    resp = requests.post(API_URL, json=payload, headers=headers, timeout=30)
    resp.raise_for_status()
    return resp.json()['translated_text']


def translate_long(text, api_key):
    """Split text that exceeds MAX_CHARS at sentence boundaries and rejoin."""
    if len(text) <= MAX_CHARS:
        return translate(text, api_key)

    # Split on '. ' or '\n' to keep semantic chunks
    sentences = []
    for part in text.replace('\n', '. ').split('. '):
        part = part.strip()
        if part:
            sentences.append(part)

    chunks, current = [], ''
    for s in sentences:
        if len(current) + len(s) + 2 <= MAX_CHARS:
            current = (current + '. ' + s).strip('. ')
        else:
            if current:
                chunks.append(current)
            current = s
    if current:
        chunks.append(current)

    parts = []
    for chunk in chunks:
        parts.append(translate(chunk, api_key))
        time.sleep(0.3)
    return '. '.join(parts)


# Column layout per sheet type:
#   OST ('On-Screen Text' / 'Translate Here'): text=col E (idx 4), translation=col F (idx 5)
#   VO  ('Narration VO'):                       text=col C (idx 2), translation=col D (idx 3)
_SHEET_CONFIG = {
    'On-Screen Text': {'text_col': 4, 'trans_col': 5},
    'Translate Here': {'text_col': 4, 'trans_col': 5},
    'Narration VO':   {'text_col': 2, 'trans_col': 3},
}


def main():
    if len(sys.argv) != 3:
        print('Usage: python auto_translate.py translations.xlsx <api_key>')
        sys.exit(1)

    xlsx_path = sys.argv[1]
    api_key = sys.argv[2]

    wb = openpyxl.load_workbook(xlsx_path)

    # Collect all rows needing translation across both sheets
    work_items = []  # (ws, row_cells, text_col, trans_col)
    for sheet_name in wb.sheetnames:
        cfg = _SHEET_CONFIG.get(sheet_name)
        if cfg is None:
            continue
        ws = wb[sheet_name]
        for row_cells in ws.iter_rows(min_row=3):
            text_cell  = row_cells[cfg['text_col']]
            trans_cell = row_cells[cfg['trans_col']]
            if (text_cell.value and str(text_cell.value).strip()
                    and not (trans_cell.value and str(trans_cell.value).strip())):
                work_items.append((sheet_name, row_cells, cfg['text_col'], cfg['trans_col']))

    total = len(work_items)
    print(f'Translating {total} rows across {len(wb.sheetnames)} sheet(s)...\n')

    done, failed = 0, 0
    for sheet_name, row_cells, text_col, trans_col in work_items:
        idx  = row_cells[0].value
        text = str(row_cells[text_col].value).strip()

        try:
            telugu = translate_long(text, api_key)
            row_cells[trans_col].value = telugu
            done += 1
            label = f'[{sheet_name}] Row {(idx or 0) + 1}'
            print(f'  [{done}/{total}] {label}: {text[:40]!r}')
            print(f'            → {telugu[:55]}')
            if done % 10 == 0:
                wb.save(xlsx_path)
            time.sleep(0.2)

        except requests.HTTPError as e:
            print(f'  [!] [{sheet_name}] Row {(idx or 0) + 1} failed ({e.response.status_code}): {text[:40]!r}')
            failed += 1
            if e.response.status_code == 429:
                print('  Rate limited — waiting 5s...')
                time.sleep(5)
        except Exception as e:
            print(f'  [!] [{sheet_name}] error: {e}')
            failed += 1

    wb.save(xlsx_path)
    print(f'\nDone. {done} translated, {failed} failed. Saved to {xlsx_path}')


if __name__ == '__main__':
    main()

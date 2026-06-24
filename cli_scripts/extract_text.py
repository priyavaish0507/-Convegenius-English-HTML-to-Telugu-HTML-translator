import sys
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup, NavigableString, Tag

SKIP_TAGS = {'script', 'style'}
PURE_NUMBER_RE = re.compile(r'^\s*-?\d+(\.\d+)?\s*$')

EMOJI_RE = re.compile(
    r'[\U0001F000-\U0001FFFF'
    r'\U00002600-\U000027FF'
    r'←-⇿'
    r'⌀-⏿'
    r'■-◿'
    r'✀-➿'
    r']+'
)


def strip_emoji(text):
    return EMOJI_RE.sub('', text).strip()


def should_extract(text):
    text = text.strip()
    if not text:
        return False
    if PURE_NUMBER_RE.match(text):
        return False
    latin_letters = re.findall(r'[a-zA-Z]', text)
    if len(latin_letters) < 3:
        return False
    return True


def walk_dom(soup):
    index = 0
    seen_attr_ids = set()

    for node in soup.descendants:
        if isinstance(node, Tag):
            if node.has_attr('data-script'):
                node_id = node.get('id', '')
                if node_id not in seen_attr_ids:
                    text = node['data-script'].strip()
                    if should_extract(text):
                        seen_attr_ids.add(node_id)
                        yield (index, 'ATTR', node_id, text)
                        index += 1

        elif isinstance(node, NavigableString) and type(node) is NavigableString:
            if any(p.name in SKIP_TAGS for p in node.parents if isinstance(p, Tag)):
                continue
            text = str(node).strip()
            if should_extract(text):
                yield (index, 'TEXT', '', text)
                index += 1


# ─── NARR / VO extraction ────────────────────────────────────────────────────

def find_narr_script_text(html_content: str) -> str | None:
    for m in re.finditer(r'<script[^>]*>(.*?)</script>', html_content,
                         re.DOTALL | re.IGNORECASE):
        if re.search(r'\bNARR\b', m.group(1)):
            return m.group(1)
    return None


def extract_narr_literal_and_offset(script_text: str) -> tuple | None:
    m = re.search(r'\bNARR\s*=\s*\{', script_text)
    if not m:
        return None

    start = m.end() - 1
    depth = 0
    in_string = False
    string_char = None
    i = start

    while i < len(script_text):
        ch = script_text[i]
        if in_string:
            if ch == '\\' and i + 1 < len(script_text):
                i += 2
                continue
            if ch == string_char:
                in_string = False
        else:
            if ch in ('"', "'", '`'):
                in_string = True
                string_char = ch
            elif ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    return script_text[start:i + 1], start
        i += 1

    return None


def _collect_value_strings(narr_literal: str) -> list:
    results = []
    i = 0
    n = len(narr_literal)

    while i < n:
        ch = narr_literal[i]

        if ch in ('"', "'", '`'):
            quote_char = ch
            i += 1
            raw_chars = []
            while i < n:
                c = narr_literal[i]
                if c == '\\' and i + 1 < n:
                    next_c = narr_literal[i + 1]
                    escapes = {
                        "'": "'", '"': '"', '\\': '\\',
                        'n': '\n', 't': '\t', 'r': '\r',
                    }
                    raw_chars.append(escapes.get(next_c, next_c))
                    i += 2
                    continue
                if c == quote_char:
                    break
                raw_chars.append(c)
                i += 1

            j = i + 1
            while j < n and narr_literal[j] in ' \t\n\r':
                j += 1
            if j < n and narr_literal[j] == ':':
                i += 1
                continue

            results.append(''.join(raw_chars))
            i += 1
            continue

        i += 1

    return results


def extract_narr_records(html_content: str) -> list:
    script_text = find_narr_script_text(html_content)
    if not script_text:
        return []
    result = extract_narr_literal_and_offset(script_text)
    if not result:
        return []
    narr_literal, _ = result
    strings = _collect_value_strings(narr_literal)
    return list(enumerate(strings))


# ─── Excel writers ───────────────────────────────────────────────────────────

def _write_ost_sheet(ws, records):
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_border = Border(
        left=Side(style='medium', color='27AE60'),
        right=thin, top=thin, bottom=thin
    )

    ws.row_dimensions[1].height = 36
    inst = Alignment(wrap_text=True, vertical='center')
    grey_font = Font(italic=True, color='555555', size=10)

    ws['E1'] = 'English text to translate (emoji removed — they will be re-added automatically).'
    ws['F1'] = '✏ Type the Telugu translation here. Write only the words — no need to add emoji.'
    for col in ['E', 'F']:
        ws[f'{col}1'].font = grey_font
        ws[f'{col}1'].alignment = inst

    ws.row_dimensions[2].height = 24
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')

    for col, label in [('E', 'Text to Translate'), ('F', 'Telugu Translation'), ('G', 'Row #')]:
        c = ws[f'{col}2']
        c.value = label
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    wrap_top = Alignment(wrap_text=True, vertical='top')

    for idx, typ, ref, text in records:
        row = idx + 3
        text_only = strip_emoji(text)

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=typ)
        ws.cell(row=row, column=3, value=ref)
        ws.cell(row=row, column=4, value=text)

        e = ws.cell(row=row, column=5, value=text_only)
        e.fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
        e.font = Font(color='1A252F', size=10)
        e.alignment = wrap_top
        e.border = border

        f = ws.cell(row=row, column=6, value='')
        f.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        f.font = Font(size=11)
        f.alignment = wrap_top
        f.border = green_border

        g = ws.cell(row=row, column=7, value=idx + 1)
        g.font = Font(color='AAAAAA', size=9)
        g.alignment = Alignment(horizontal='center', vertical='top')

        ws.row_dimensions[row].height = max(18, min(80, 15 + (len(text) // 50) * 14))

    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].hidden = True
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 7
    ws.freeze_panes = 'E3'


def _write_vo_sheet(ws, vo_records):
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    orange_border = Border(
        left=Side(style='medium', color='E67E22'),
        right=thin, top=thin, bottom=thin
    )

    ws.row_dimensions[1].height = 36
    grey_font = Font(italic=True, color='555555', size=10)
    inst = Alignment(wrap_text=True, vertical='center')
    ws['C1'] = 'Narration text from the NARR object (voice-over). Each row is one spoken utterance.'
    ws['D1'] = '✏ Type the Telugu narration here. These strings will be spoken aloud by the browser.'
    for col in ['C', 'D']:
        ws[f'{col}1'].font = grey_font
        ws[f'{col}1'].alignment = inst

    ws.row_dimensions[2].height = 24
    header_fill = PatternFill(start_color='B7470A', end_color='B7470A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    for col, label in [('C', 'Narration to Translate'), ('D', 'Telugu Narration'), ('E', 'Row #')]:
        c = ws[f'{col}2']
        c.value = label
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    wrap_top = Alignment(wrap_text=True, vertical='top')
    for idx, text in vo_records:
        row = idx + 3
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=text)

        c = ws.cell(row=row, column=3, value=text)
        c.fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
        c.font = Font(color='1A252F', size=10)
        c.alignment = wrap_top
        c.border = border

        d = ws.cell(row=row, column=4, value='')
        d.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        d.font = Font(size=11)
        d.alignment = wrap_top
        d.border = orange_border

        e = ws.cell(row=row, column=5, value=idx + 1)
        e.font = Font(color='AAAAAA', size=9)
        e.alignment = Alignment(horizontal='center', vertical='top')

        ws.row_dimensions[row].height = max(18, min(80, 15 + (len(text) // 50) * 14))

    ws.column_dimensions['A'].hidden = True
    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 7
    ws.freeze_panes = 'C3'


def write_excel(ost_records, vo_records, output_path):
    wb = openpyxl.Workbook()

    ws_ost = wb.active
    ws_ost.title = 'On-Screen Text'
    ws_ost.sheet_properties.tabColor = '1A5276'
    _write_ost_sheet(ws_ost, ost_records)

    if vo_records:
        ws_vo = wb.create_sheet('Narration VO')
        ws_vo.sheet_properties.tabColor = 'E67E22'
        _write_vo_sheet(ws_vo, vo_records)

    wb.save(output_path)


def main():
    if len(sys.argv) != 3:
        print('Usage: python extract_text.py "<input.html>" <output.xlsx>')
        sys.exit(1)

    html_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    with open(html_path, encoding='utf-8') as f:
        html = f.read()

    soup = BeautifulSoup(html, 'html.parser')
    ost_records = list(walk_dom(soup))
    vo_records = extract_narr_records(html)

    write_excel(ost_records, vo_records, xlsx_path)
    print(f'Extracted {len(ost_records)} OST + {len(vo_records)} VO items to {xlsx_path}')


if __name__ == '__main__':
    main()

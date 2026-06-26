"""
Core extract/apply logic.
All functions operate on strings/BytesIO — no file system access.
"""
import re
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup, NavigableString, Tag

SKIP_TAGS = {'script', 'style'}
PURE_NUMBER_RE = re.compile(r'^\s*-?\d+(\.\d+)?\s*$')

EMOJI_RE = re.compile(
    r'[\U0001F000-\U0001FFFF'
    r'\U00002600-\U000027FF'
    r'←-⇿'
    r'⌀-⏿'
    r'■-◿'
    r'✀-➿'
    r']+'
)
_LEAD_PAT = re.compile(
    r'^([\s\U0001F000-\U0001FFFF\U00002600-\U000027FF←-⇿⌀-⏿■-◿✀-➿]*)'
)
_TRAIL_PAT = re.compile(
    r'([\s\U0001F000-\U0001FFFF\U00002600-\U000027FF←-⇿⌀-⏿■-◿✀-➿]*)$'
)


# ─── OST (on-screen text) helpers ────────────────────────────────────────────

def strip_emoji(text):
    return EMOJI_RE.sub('', text).strip()


def merge_emoji(original, translated):
    lead_m = _LEAD_PAT.match(original)
    trail_m = _TRAIL_PAT.search(original)
    leading = lead_m.group().strip() if lead_m and EMOJI_RE.search(lead_m.group()) else ''
    trailing = trail_m.group().strip() if trail_m and EMOJI_RE.search(trail_m.group()) else ''
    result = translated.strip()
    if leading:
        result = leading + ' ' + result
    if trailing and trailing != leading:
        result = result + ' ' + trailing
    return result


def should_extract(text):
    text = text.strip()
    if not text:
        return False
    if PURE_NUMBER_RE.match(text):
        return False
    if len(re.findall(r'[a-zA-Z]', text)) < 2:
        return False
    return True


def walk_dom(soup):
    index = 0
    seen_attr_ids = set()
    for node in soup.descendants:
        if isinstance(node, Tag):
            if node.has_attr('data-script'):
                node_id = node.get('id', '')
                if node_id not in seen_attr_ids:
                    text = node['data-script'].strip()
                    if should_extract(text):
                        seen_attr_ids.add(node_id)
                        yield (index, 'ATTR', node_id, text)
                        index += 1
        elif isinstance(node, NavigableString) and type(node) is NavigableString:
            if any(p.name in SKIP_TAGS for p in node.parents if isinstance(p, Tag)):
                continue
            text = str(node).strip()
            if should_extract(text):
                yield (index, 'TEXT', '', text)
                index += 1


# ─── NARR / VO helpers ───────────────────────────────────────────────────────

def find_narr_script_text(html_content: str) -> str | None:
    """Return the raw text content of the <script> tag containing NARR, or None."""
    for m in re.finditer(r'<script[^>]*>(.*?)</script>', html_content,
                         re.DOTALL | re.IGNORECASE):
        if re.search(r'\bNARR\b', m.group(1)):
            return m.group(1)
    return None


def _find_narr_script_offset(html_content: str) -> tuple[str, int] | None:
    """Return (script_content, start_offset_in_html) for the <script> containing NARR."""
    for m in re.finditer(r'<script[^>]*>(.*?)</script>', html_content,
                         re.DOTALL | re.IGNORECASE):
        if re.search(r'\bNARR\b', m.group(1)):
            return m.group(1), m.start(1)
    return None


def extract_narr_literal_and_offset(script_text: str) -> tuple[str, int] | None:
    """
    Find 'NARR = {...}' and return (literal, start_offset_in_script_text).
    Uses brace-counting with string-aware parsing to locate the matching '}'.
    """
    m = re.search(r'\bNARR\s*=\s*\{', script_text)
    if not m:
        return None

    start = m.end() - 1   # position of the opening '{'
    depth = 0
    in_string = False
    string_char = None
    i = start

    while i < len(script_text):
        ch = script_text[i]
        if in_string:
            if ch == '\\' and i + 1 < len(script_text):
                i += 2
                continue
            if ch == string_char:
                in_string = False
        else:
            if ch in ('"', "'", '`'):
                in_string = True
                string_char = ch
            elif ch == '{':
                depth += 1
            elif ch == '}':
                depth -= 1
                if depth == 0:
                    return script_text[start:i + 1], start
        i += 1

    return None


def _collect_value_strings(narr_literal: str) -> list[str]:
    """
    Extract all value-position quoted strings from a JS object literal.
    Structure-agnostic: a string is considered a key only when a ':' immediately
    follows it (after optional whitespace) within an object context.
    Returns decoded string values in source order.
    """
    results = []
    i = 0
    n = len(narr_literal)

    while i < n:
        ch = narr_literal[i]

        if ch in ('"', "'", '`'):
            quote_char = ch
            i += 1
            raw_chars = []
            while i < n:
                c = narr_literal[i]
                if c == '\\' and i + 1 < n:
                    next_c = narr_literal[i + 1]
                    escapes = {
                        "'": "'", '"': '"', '\\': '\\',
                        'n': '\n', 't': '\t', 'r': '\r',
                    }
                    raw_chars.append(escapes.get(next_c, next_c))
                    i += 2
                    continue
                if c == quote_char:
                    break
                raw_chars.append(c)
                i += 1

            # Look ahead: if ':' follows (with optional whitespace) it's a property key
            j = i + 1
            while j < n and narr_literal[j] in ' \t\n\r':
                j += 1
            if j < n and narr_literal[j] == ':':
                i += 1  # skip closing quote, this is a key
                continue

            results.append(''.join(raw_chars))
            i += 1  # skip closing quote
            continue

        i += 1

    return results


def extract_narr_records(html_content: str) -> list[tuple[int, str]]:
    """
    Return [(sequential_index, string_value), ...] for all narration strings
    inside the NARR object. Completely structure-agnostic.
    Returns [] if no NARR found.
    """
    script_text = find_narr_script_text(html_content)
    if not script_text:
        return []
    result = extract_narr_literal_and_offset(script_text)
    if not result:
        return []
    narr_literal, _ = result
    strings = _collect_value_strings(narr_literal)
    return list(enumerate(strings))


def _js_escape(text: str, quote_char: str) -> str:
    """Escape text for safe embedding inside a JS string with the given quote style."""
    text = text.replace('\\', '\\\\')
    text = text.replace(quote_char, '\\' + quote_char)
    return text


def _rebuild_narr_literal(
    narr_literal: str,
    translations: dict,
    english_check: dict,
) -> tuple[str, int]:
    """
    Walk narr_literal and substitute value-position strings with translations.
    Returns (new_literal, applied_count).
    """
    output = []
    value_idx = 0
    applied = 0
    i = 0
    n = len(narr_literal)

    while i < n:
        ch = narr_literal[i]

        if ch in ('"', "'", '`'):
            quote_char = ch
            end = i + 1
            while end < n:
                c = narr_literal[end]
                if c == '\\' and end + 1 < n:
                    end += 2
                    continue
                if c == quote_char:
                    break
                end += 1

            # Check if this is a property key
            j = end + 1
            while j < n and narr_literal[j] in ' \t\n\r':
                j += 1
            is_key = j < n and narr_literal[j] == ':'

            if is_key:
                output.append(narr_literal[i:end + 1])
            else:
                if value_idx in translations:
                    telugu = translations[value_idx]
                    escaped = _js_escape(telugu, quote_char)
                    output.append(f'{quote_char}{escaped}{quote_char}')
                    applied += 1
                else:
                    output.append(narr_literal[i:end + 1])
                value_idx += 1

            i = end + 1
            continue

        output.append(ch)
        i += 1

    return ''.join(output), applied


def apply_narr_translations(
    html_content: str,
    vo_translations: dict,
    vo_english_check: dict,
) -> tuple[str, int, int]:
    """
    Replace NARR strings in the raw HTML string (not through BeautifulSoup,
    to avoid script-content encoding issues).
    Returns (new_html, applied_count, total_count).
    """
    result = _find_narr_script_offset(html_content)
    if not result:
        return html_content, 0, 0
    script_text, script_offset = result

    narr_result = extract_narr_literal_and_offset(script_text)
    if not narr_result:
        return html_content, 0, 0
    narr_literal, narr_start_in_script = narr_result

    new_narr, applied = _rebuild_narr_literal(narr_literal, vo_translations, vo_english_check)

    abs_start = script_offset + narr_start_in_script
    abs_end = abs_start + len(narr_literal)
    new_html = html_content[:abs_start] + new_narr + html_content[abs_end:]

    total = len(vo_english_check)
    return new_html, applied, total


# ─── playWelcome bug fix ─────────────────────────────────────────────────────

def fix_play_welcome(html_content: str) -> str:
    """
    Fix the playWelcome first-load audio bug: some HTML templates join all
    NARR[0].parts into a single string before calling speakOne(text, ...).
    That joined string never matches any key in AUDIO_CLIPS (which stores
    individual part strings), so the welcome audio silently falls back to
    browser TTS on first load even though clips exist.

    replaySlide already uses speakChain(parts, ...) correctly — this makes
    playWelcome consistent with it.

    No-op if the pattern is absent (template already correct or different).
    """
    # Pattern 1: remove the join line and fix the empty-check condition.
    # Handles both pretty-printed (spaces around = / ||) and minified forms.
    # Before: const text = parts.join("  ");\nif (!text ||   (pretty)
    # Before: const text=parts.join('  ');\nif(!text||        (minified)
    # After:  if(!parts.length||
    fixed, n = re.subn(
        r'const text\s*=\s*parts\.join\(["\']  ["\']\)\s*;[ \t]*\n[ \t]*if\s*\(\s*!text\s*\|\|',
        'if(!parts.length||',
        html_content,
    )
    if n == 0:
        return html_content  # pattern not present — nothing to do

    # Pattern 2: replace speakOne(text, ...) with speakChain(parts, ...).
    # Handles both pretty-printed and minified whitespace around ( ) and commas.
    fixed, _ = re.subn(
        r'\bspeakOne\(\s*text\s*,\s*function\s*\(\s*\)\s*\{',
        'speakChain(parts,function(){',
        fixed,
    )
    return fixed


# ─── Voice language patcher ──────────────────────────────────────────────────

# Sentinel that both update_voice_language and inject_audio_clips embed, so
# whichever runs first the second one is a no-op (prevents double-patching).
_TELUGU_PATCH_GUARD = '/* __telugu-tts-patch__ */'

# Injected when there are no pre-generated audio clips (text-only path).
# Wraps speechSynthesis.speak at the lowest level so it works regardless of
# what the HTML calls its internal TTS helper (say, speakOne, speak, …).
# Clears any non-Telugu voice that rankVoice() may have auto-selected, then
# forces lang='te-IN' so the browser picks the best available Telugu voice.
_SPEAK_LANG_PATCH = (
    '\n<script>\n'
    f'{_TELUGU_PATCH_GUARD}\n'
    '(function () {\n'
    '  if (typeof speechSynthesis === "undefined") return;\n'
    '  const _s = speechSynthesis.speak.bind(speechSynthesis);\n'
    '  speechSynthesis.speak = function (u) {\n'
    '    if (u.voice && !u.voice.lang.toLowerCase().startsWith("te")) u.voice = null;\n'
    '    u.lang = "te-IN";\n'
    '    _s(u);\n'
    '  };\n'
    '})();\n'
    '</script>'
)


def update_voice_language(html_content: str) -> str:
    """
    Intercept speechSynthesis.speak to force te-IN on every utterance.
    Works regardless of the HTML's internal function names or rankVoice logic.
    No-op if already patched (guard comment present).
    """
    if _TELUGU_PATCH_GUARD in html_content:
        return html_content
    close_body = html_content.rfind('</body>')
    if close_body != -1:
        return html_content[:close_body] + _SPEAK_LANG_PATCH + '\n' + html_content[close_body:]
    return html_content + _SPEAK_LANG_PATCH


def inject_audio_clips(html_content: str, audio_clips: dict) -> str:
    """
    Inject AUDIO_CLIPS and wrap speechSynthesis.speak to play pre-generated
    MP3s instead of browser TTS. Falls back to te-IN TTS for any clip that
    wasn't generated. Works regardless of what the HTML calls its TTS helper.
    Inserts a <script> tag just before </body>.
    """
    if not audio_clips:
        return html_content

    clips_json = json.dumps(audio_clips, ensure_ascii=False, indent=2)
    patch_script = (
        '\n<script>\n'
        f'{_TELUGU_PATCH_GUARD}\n'
        f'const AUDIO_CLIPS = {clips_json};\n'
        '(function () {\n'
        '  if (typeof speechSynthesis === "undefined") return;\n'
        '  const _s = speechSynthesis.speak.bind(speechSynthesis);\n'
        '  let _cur = null;\n'
        '  speechSynthesis.speak = function (u) {\n'
        '    if (_cur) { _cur.pause(); _cur.currentTime = 0; _cur = null; }\n'
        '    const clip = AUDIO_CLIPS[u.text];\n'
        '    if (clip) {\n'
        '      const a = new Audio(clip);\n'
        '      _cur = a;\n'
        '      setTimeout(function () { if (u.onstart) u.onstart({}); }, 0);\n'
        '      a.onended = function () { _cur = null; if (u.onend) u.onend({}); };\n'
        '      a.onerror = function () { _cur = null; u.voice = null; u.lang = "te-IN"; _s(u); };\n'
        '      a.play().catch(function () { _cur = null; u.voice = null; u.lang = "te-IN"; _s(u); });\n'
        '      return;\n'
        '    }\n'
        '    if (u.voice && !u.voice.lang.toLowerCase().startsWith("te")) u.voice = null;\n'
        '    u.lang = "te-IN";\n'
        '    _s(u);\n'
        '  };\n'
        '})();\n'
        '</script>'
    )
    close_body = html_content.rfind('</body>')
    if close_body != -1:
        return html_content[:close_body] + patch_script + '\n' + html_content[close_body:]
    return html_content + patch_script


# ─── Excel sheet writers ─────────────────────────────────────────────────────

def _write_ost_sheet(ws, records):
    """Write OST records into ws (already created by the caller)."""
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_border = Border(
        left=Side(style='medium', color='27AE60'),
        right=thin, top=thin, bottom=thin
    )

    ws.row_dimensions[1].height = 36
    grey_font = Font(italic=True, color='555555', size=10)
    inst = Alignment(wrap_text=True, vertical='center')
    ws['E1'] = 'English text to translate (emoji removed — they will be re-added automatically).'
    ws['F1'] = '✏ Type the Telugu translation here. Write only the words — no need to add emoji.'
    for col in ['E', 'F']:
        ws[f'{col}1'].font = grey_font
        ws[f'{col}1'].alignment = inst

    ws.row_dimensions[2].height = 24
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    for col, label in [('E', 'Text to Translate'), ('F', 'Telugu Translation'), ('G', 'Row #')]:
        c = ws[f'{col}2']
        c.value = label
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    wrap_top = Alignment(wrap_text=True, vertical='top')
    for idx, typ, ref, text in records:
        row = idx + 3
        text_only = strip_emoji(text)
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=typ)
        ws.cell(row=row, column=3, value=ref)
        ws.cell(row=row, column=4, value=text)

        e = ws.cell(row=row, column=5, value=text_only)
        e.fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
        e.font = Font(color='1A252F', size=10)
        e.alignment = wrap_top
        e.border = border

        f = ws.cell(row=row, column=6, value='')
        f.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        f.font = Font(size=11)
        f.alignment = wrap_top
        f.border = green_border

        g = ws.cell(row=row, column=7, value=idx + 1)
        g.font = Font(color='AAAAAA', size=9)
        g.alignment = Alignment(horizontal='center', vertical='top')

        ws.row_dimensions[row].height = max(18, min(80, 15 + (len(text) // 50) * 14))

    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].hidden = True
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 7
    ws.freeze_panes = 'E3'


def _write_vo_sheet(ws, vo_records):
    """Write VO narration records into ws (already created by the caller)."""
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    orange_border = Border(
        left=Side(style='medium', color='E67E22'),
        right=thin, top=thin, bottom=thin
    )

    ws.row_dimensions[1].height = 36
    grey_font = Font(italic=True, color='555555', size=10)
    inst = Alignment(wrap_text=True, vertical='center')
    ws['C1'] = 'Narration text from the NARR object (voice-over). Each row is one spoken utterance.'
    ws['D1'] = '✏ Type the Telugu narration here. These strings will be spoken aloud by the browser.'
    for col in ['C', 'D']:
        ws[f'{col}1'].font = grey_font
        ws[f'{col}1'].alignment = inst

    ws.row_dimensions[2].height = 24
    header_fill = PatternFill(start_color='B7470A', end_color='B7470A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    center = Alignment(horizontal='center', vertical='center')
    for col, label in [('C', 'Narration to Translate'), ('D', 'Telugu Narration'), ('E', 'Row #')]:
        c = ws[f'{col}2']
        c.value = label
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    wrap_top = Alignment(wrap_text=True, vertical='top')
    for idx, text in vo_records:
        row = idx + 3
        ws.cell(row=row, column=1, value=idx)    # A hidden: sequential index
        ws.cell(row=row, column=2, value=text)   # B hidden: original English

        c = ws.cell(row=row, column=3, value=text)   # C visible: English display
        c.fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
        c.font = Font(color='1A252F', size=10)
        c.alignment = wrap_top
        c.border = border

        d = ws.cell(row=row, column=4, value='')     # D visible: Telugu input
        d.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        d.font = Font(size=11)
        d.alignment = wrap_top
        d.border = orange_border

        e = ws.cell(row=row, column=5, value=idx + 1)  # E visible: row number
        e.font = Font(color='AAAAAA', size=9)
        e.alignment = Alignment(horizontal='center', vertical='top')

        ws.row_dimensions[row].height = max(18, min(80, 15 + (len(text) // 50) * 14))

    ws.column_dimensions['A'].hidden = True
    ws.column_dimensions['B'].hidden = True
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 7
    ws.freeze_panes = 'C3'


# ─── Public API ──────────────────────────────────────────────────────────────

def extract_to_excel(html_content: str) -> bytes:
    """
    Parse HTML, extract OST and VO narration text.
    Returns xlsx bytes with two sheets: 'On-Screen Text' and 'Narration VO'.
    """
    soup = BeautifulSoup(html_content, 'html.parser')
    ost_records = list(walk_dom(soup))
    vo_records = extract_narr_records(html_content)

    wb = openpyxl.Workbook()

    ws_ost = wb.active
    ws_ost.title = 'On-Screen Text'
    ws_ost.sheet_properties.tabColor = '1A5276'
    _write_ost_sheet(ws_ost, ost_records)

    if vo_records:
        ws_vo = wb.create_sheet('Narration VO')
        ws_vo.sheet_properties.tabColor = 'E67E22'
        _write_vo_sheet(ws_vo, vo_records)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_translations_from_bytes(xlsx_bytes: bytes):
    """
    Read filled Excel, return (ost_trans, ost_check, vo_trans, vo_check).
    Backward compatible: old single-sheet Excel returns empty VO dicts.

    ost_trans / ost_check: {int_idx: str}
    vo_trans  / vo_check:  {int_idx: str}  (sequential index into NARR value strings)
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    # OST sheet — support both legacy name and new name
    if 'On-Screen Text' in wb.sheetnames:
        ost_ws = wb['On-Screen Text']
    elif 'Translate Here' in wb.sheetnames:
        ost_ws = wb['Translate Here']
    else:
        ost_ws = wb.active

    ost_trans, ost_check = {}, {}
    for row in ost_ws.iter_rows(min_row=3, values_only=True):
        if row[0] is None:
            continue
        idx = int(row[0])
        ost_check[idx] = str(row[3] or '').strip()
        telugu = row[5]
        if telugu and str(telugu).strip():
            ost_trans[idx] = str(telugu).strip()

    vo_trans, vo_check = {}, {}
    if 'Narration VO' in wb.sheetnames:
        vo_ws = wb['Narration VO']
        for row in vo_ws.iter_rows(min_row=3, values_only=True):
            if row[0] is None:
                continue
            idx      = int(row[0])           # col A: sequential index
            original = str(row[1] or '').strip()  # col B: original English
            telugu   = row[3]                # col D: translation
            vo_check[idx] = original
            if telugu and str(telugu).strip():
                vo_trans[idx] = str(telugu).strip()

    return ost_trans, ost_check, vo_trans, vo_check


def apply_translations(html_content: str, translations: dict, english_check: dict) -> tuple[str, int, int]:
    """Apply OST translations to HTML DOM. Returns (new_html, applied_count, total_count)."""
    soup = BeautifulSoup(html_content, 'html.parser')

    replacements = []
    index = 0
    seen_attr_ids = set()

    for node in soup.descendants:
        if isinstance(node, Tag):
            if node.has_attr('data-script'):
                node_id = node.get('id', '')
                if node_id not in seen_attr_ids:
                    text = node['data-script'].strip()
                    if should_extract(text):
                        seen_attr_ids.add(node_id)
                        if index in translations:
                            final = merge_emoji(english_check.get(index, text), translations[index])
                            replacements.append((node, final, 'ATTR'))
                        index += 1
        elif isinstance(node, NavigableString) and type(node) is NavigableString:
            if any(p.name in SKIP_TAGS for p in node.parents if isinstance(p, Tag)):
                continue
            text = str(node).strip()
            if should_extract(text):
                if index in translations:
                    final = merge_emoji(english_check.get(index, text), translations[index])
                    replacements.append((node, final, 'TEXT'))
                index += 1

    total = index

    for node, new_text, node_type in replacements:
        if node_type == 'ATTR':
            node['data-script'] = new_text
        else:
            orig = str(node)
            lead = orig[:len(orig) - len(orig.lstrip())]
            trail = orig[len(orig.rstrip()):]
            node.replace_with(NavigableString(lead + new_text + trail))

    return str(soup), len(replacements), total
